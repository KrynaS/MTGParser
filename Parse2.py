import json
from xlwt import Workbook
from openpyxl import load_workbook

def printkolory(kolory):
	str = ""
	for y in range(len(kolory)):
		if y+1 != len(kolory):
			str = str + kolory[y] + "/"
		else:
			str = str + kolory[y]
	return str

def printRarity(rarity):
	str = ''
	first = True
	if 'common' in rarity:
		str += 'C'
		first = False
	if 'uncommon' in rarity:
		str += 'U' if first else '/U'
		first = False
	if 'rare' in rarity:
		str += 'R' if first else '/R'
		first = False
	if 'mythic' in rarity:
		str += 'M' if first else '/M'
		first = False
	return str

class Karta:

	def __init__(self, name, kolor, rarity, ilosc, misc):
		self.name = name
		self.kolor = kolor
		self.rarity = rarity
		self.ilosc = ilosc
		self.misc = misc

	def isLowestRarity(self, rarity):

		if self.isThisRarity(rarity):
			if rarity == 'common':
				return True
			if rarity == 'uncommon' and 'common' not in self.rarity :
				return True
			if any(x in rarity for x in ['rare', 'mythic']) and 'common' not in self.rarity and 'uncommon' not in self.rarity:
				return True
			
		return False

	def isThisRarity(self, rarity):
		return any(x in rarity for x in self.rarity)

class Karty:
	
	def __init__(self, dic):
		self.fullList = dic.values()

	# def __init__(self, dic1, dic2):
	# 	for card in dic2:
	# 		if dic1[card.name] == None:
	# 			dic1[card.name] = card

	def getByExactColorAndRarity(self, color, rarity):
		result = list()

		for karta in self.fullList:
			if color in karta.kolor and len(karta.kolor) == 1 and karta.isLowestRarity(rarity):
				result.append(karta)
		result.sort(key=lambda x: x.name, reverse=False)
		return result

	def getMultiColorsByRarity(self, rarity):
		result = list()

		for karta in self.fullList:
			if len(karta.kolor) > 1 and karta.isLowestRarity(rarity):
				result.append(karta)
		result.sort(key=lambda x: x.name, reverse=False)
		return result

	def getNoColorByRarity(self, rarity):
		result = list()

		for karta in self.fullList:
			if len(karta.kolor) == 0 and karta.isLowestRarity(rarity):
				result.append(karta)
		result.sort(key=lambda x: x.name, reverse=False)
		return result	

f = open('AllPrintings.json', encoding='utf-8')
data = json.load(f)
dic = dict()
data = data['data']

for mset in data.values():
	if mset.get('isPartialPreview', None) == True:
		continue
	if mset.get("name") == 'Renaissance':
		continue
	for card in mset['cards']:
		if card['name'] == "Archangel Avacyn // Avacyn, the Purifier":
			stop = 0
		if card['name'] == "Wear // Tear":
			stop = 0
		if card.get('isOnlineOnly', None) == True:
			continue
		if dic.get(card['name']) != None:
			if card.get('side') != None and not set(dic.get(card['name']).kolor) >= set(card['colors']) and card.get("manaCost") != None:
				dic.get(card['name']).kolor += card['colors']

			if card['rarity'] not in dic[card['name']].rarity:
				# if card['name'] == 'Serra Angel':
				# 	stop = 0
				dic[card['name']].rarity.append(card['rarity'])
		else:
			# if card['legalities'].get('pioneer') != None and card['legalities'].get('pioneer') == 'Legal':
			if card.get('text', None) != None and "Devoid" in card['text']:
				if card.get('manaCost', None) != None:
					if '{W}' in card['manaCost'] and 'W' not in card['colors']:
						card['colors'].append('W')
					if '{U}' in card['manaCost'] and 'U' not in card['colors']:
						card['colors'].append('U')
					if '{B}' in card['manaCost'] and 'B' not in card['colors']:
						card['colors'].append('B')
					if '{R}' in card['manaCost'] and 'R' not in card['colors']:
						card['colors'].append('R')
					if '{G}' in card['manaCost'] and 'G' not in card['colors']:
						card['colors'].append('G')
			karta = Karta(card['name'], card['colors'], [card['rarity']], 0, "")
			dic[karta.name] = karta

wb1 = load_workbook('karty.xlsx')
ws1 = wb1.active

for row in ws1.iter_rows(max_col=5, values_only=True):
	if type(row[3]) == float and row[3] > 0:
		if dic.get(row[0]) != None:
			dic.get(row[0]).ilosc = row[3]
			dic.get(row[0]).misc = row[4]

	
karty = Karty(dic)

wb = Workbook()
sheet = wb.add_sheet('Karty')

kolory = ['W', 'U', 'B', 'R', 'G', ]
rarity = ['common', 'uncommon', ['rare', 'mythic']]

i = 0
for k in kolory:
	for r in rarity:
		for x in karty.getByExactColorAndRarity(k, r):
			#f.write(lista[x].name  + "." + printkolory(lista[x].kolor)  + "." +  lista[x].rarity  + "\n")
			sheet.write(i, 0, x.name)
			sheet.write(i, 1, x.kolor)
			sheet.write(i, 2, printRarity(x.rarity))
			sheet.write(i, 3, x.ilosc)
			sheet.write(i, 4, x.misc)
			i+=1

for r in rarity:
	for x in karty.getMultiColorsByRarity(r):
		sheet.write(i, 0, x.name)
		sheet.write(i, 1, x.kolor)
		sheet.write(i, 2, printRarity(x.rarity))
		sheet.write(i, 3, x.ilosc)
		sheet.write(i, 4, x.misc)
		i+=1

for r in rarity:
	for x in karty.getNoColorByRarity(r):
		sheet.write(i, 0, x.name)
		sheet.write(i, 1, x.kolor)
		sheet.write(i, 2, printRarity(x.rarity))
		sheet.write(i, 3, x.ilosc)
		sheet.write(i, 4, x.misc)
		i+=1

wb.save('kartynew.xls')